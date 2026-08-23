
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Blackbox Testing"

# ─── Color Palette ───────────────────────────────────────────────────────────
DARK_BROWN   = "3E1C00"   # header bg
MEDIUM_BROWN = "6F3D0D"   # sub-header / feature label
LIGHT_CREAM  = "FFF8F0"   # alternating row 1
WHITE_ROW    = "FFFFFF"   # alternating row 2
ACCENT_GOLD  = "C8860A"   # column header bg
BERHASIL_BG  = "D6F4E3"   # status passed
GAGAL_BG     = "FAD7D7"   # status failed
BORDER_COLOR = "BFAC98"
TEXT_DARK    = "1A0A00"
TEXT_WHITE   = "FFFFFF"
TEXT_GOLD    = "C8860A"

# ─── Helper: make a thin border ──────────────────────────────────────────────
def thin_border(color=BORDER_COLOR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    tk = Side(style="medium", color=DARK_BROWN)
    tn = Side(style="thin",   color=BORDER_COLOR)
    return Border(left=tk, right=tk, top=tn, bottom=tn)

# ─── Column widths (A..F) ───────────────────────────────────────────────────
col_widths = {
    "A": 5,    # No
    "B": 22,   # Fitur Aplikasi
    "C": 30,   # Deskripsi Pengujian
    "D": 42,   # Kondisi Pengujian
    "E": 48,   # Hasil yang Diharapkan
    "F": 13,   # Status
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# ─── ROW 1: Main Title ───────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "Tabel Hasil Pengujian Blackbox Testing — Aplikasi Skena Coffee"
title_cell.font      = Font(name="Calibri", bold=True, size=14, color=TEXT_WHITE)
title_cell.fill      = PatternFill("solid", fgColor=DARK_BROWN)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
title_cell.border    = thin_border(DARK_BROWN)
ws.row_dimensions[1].height = 30

# ─── ROW 2: Column Headers ───────────────────────────────────────────────────
headers = ["No", "Fitur Aplikasi", "Deskripsi Pengujian",
           "Kondisi Pengujian", "Hasil yang Diharapkan", "Status"]
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.font      = Font(name="Calibri", bold=True, size=10, color=TEXT_WHITE)
    cell.fill      = PatternFill("solid", fgColor=ACCENT_GOLD)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[2].height = 28

# ─── Test Data ───────────────────────────────────────────────────────────────
# Each entry: (fitur, deskripsi, kondisi, hasil_diharapkan, status)
test_data = [
    # ── HALAMAN UTAMA (HOME) ─────────────────────────────────────────────────
    ("Halaman Utama (Home)",
     "Menampilkan halaman utama aplikasi",
     "Pengguna mengakses URL '/' tanpa login",
     "Sistem menampilkan halaman landing page Skena Coffee beserta daftar menu best seller (4 item teratas berdasarkan rating)",
     "Berhasil"),

    ("Halaman Utama (Home)",
     "Menampilkan menu best seller",
     "Terdapat minimal 4 menu yang tersedia (is_available = true) di database",
     "Sistem menampilkan 4 menu dengan rating tertinggi secara otomatis pada section best seller",
     "Berhasil"),

    # ── HALAMAN MENU ─────────────────────────────────────────────────────────
    ("Halaman Menu",
     "Menampilkan daftar menu berdasarkan kategori",
     "Pengguna membuka halaman '/menu'",
     "Sistem menampilkan seluruh menu yang tersedia, dikelompokkan berdasarkan kategori, diurutkan sesuai sort_order",
     "Berhasil"),

    ("Halaman Menu",
     "Filter menu per kategori",
     "Pengguna memilih salah satu tab kategori (misal: Kopi Panas)",
     "Sistem hanya menampilkan menu yang termasuk dalam kategori yang dipilih tanpa reload halaman",
     "Berhasil"),

    ("Halaman Menu",
     "Menampilkan detail menu",
     "Pengguna menekan salah satu kartu menu",
     "Sistem mengarahkan pengguna ke halaman '/menu/{id}' yang menampilkan nama, deskripsi, harga, dan menu rekomendasi terkait",
     "Berhasil"),

    ("Halaman Menu",
     "Menu tidak ditemukan (404)",
     "Pengguna mengakses '/menu/9999' dengan ID yang tidak ada di database",
     "Sistem menampilkan halaman error 404 (Not Found)",
     "Berhasil"),

    # ── SCAN QR CODE ─────────────────────────────────────────────────────────
    ("Scan QR Code",
     "Redirect setelah scan QR meja",
     "Pengguna melakukan scan QR Code meja (misal: '/scan/5')",
     "Sistem melakukan redirect ke halaman '/menu?table=5' sehingga nomor meja tersimpan dan dikenali",
     "Berhasil"),

    # ── KERANJANG (CART) ─────────────────────────────────────────────────────
    ("Keranjang (Cart)",
     "Menambahkan menu ke keranjang",
     "Pengguna memilih menu dan menekan tombol 'Tambah ke Keranjang'",
     "Item ditambahkan ke keranjang (localStorage), badge jumlah item pada ikon keranjang bertambah",
     "Berhasil"),

    ("Keranjang (Cart)",
     "Mengubah kuantitas item di keranjang",
     "Pengguna menekan tombol '+' atau '−' pada item di halaman keranjang",
     "Kuantitas item berubah sesuai tindakan, subtotal dan total harga diperbarui secara real-time",
     "Berhasil"),

    ("Keranjang (Cart)",
     "Menghapus item dari keranjang",
     "Pengguna menekan tombol hapus (ikon tong sampah) pada salah satu item",
     "Item dihapus dari keranjang, total harga diperbarui. Jika keranjang kosong, tampil pesan 'Keranjang Kosong'",
     "Berhasil"),

    ("Keranjang (Cart)",
     "Keranjang kosong saat checkout",
     "Pengguna mengakses halaman checkout tanpa ada item di keranjang",
     "Sistem menampilkan pesan keranjang kosong dan menonaktifkan tombol Checkout",
     "Berhasil"),

    # ── CHECKOUT ─────────────────────────────────────────────────────────────
    ("Checkout",
     "Validasi form checkout — data wajib kosong",
     "Pengguna menekan tombol 'Bayar Sekarang' tanpa mengisi nama pelanggan atau nomor meja",
     "Sistem menampilkan pesan validasi: 'Nama pelanggan wajib diisi' dan 'Nomor meja wajib diisi', transaksi tidak diproses",
     "Berhasil"),

    ("Checkout",
     "Menghitung subtotal, pajak (10%), dan total",
     "Pengguna memiliki item di keranjang dengan subtotal Rp 50.000",
     "Sistem menampilkan: Subtotal = Rp 50.000, Pajak 10% = Rp 5.000, Total = Rp 55.000",
     "Berhasil"),

    ("Checkout",
     "Membuat transaksi QRIS (createQris)",
     "Pengguna mengisi nama, nomor meja, dan menekan 'Bayar dengan QRIS'",
     "Sistem memanggil API Midtrans Core, menghasilkan QR code URL, menyimpan order ke database dengan status 'pending', dan menampilkan QR code kepada pengguna",
     "Berhasil"),

    ("Checkout",
     "Membuat transaksi Snap (createToken) sebagai fallback",
     "Pengguna memilih metode pembayaran selain QRIS melalui Midtrans Snap",
     "Sistem memanggil Midtrans Snap API, mendapatkan token dan redirect URL, menyimpan order dengan status 'pending'",
     "Berhasil"),

    ("Checkout",
     "Validasi input checkout — items kosong",
     "Request POST ke '/order/create-qris' dikirim tanpa field 'items'",
     "Sistem mengembalikan response error 422 Unprocessable Entity dengan pesan validasi 'items wajib diisi'",
     "Berhasil"),

    # ── STATUS PESANAN ────────────────────────────────────────────────────────
    ("Status Pesanan",
     "Menampilkan halaman status pesanan",
     "Pengguna mengakses '/order/status?order_id=SKENA-XXXXXX-YYYYYY'",
     "Sistem menampilkan detail pesanan (nama pelanggan, daftar item, total, status) sesuai order_id yang diberikan",
     "Berhasil"),

    ("Status Pesanan",
     "Menampilkan status pesanan tidak ditemukan",
     "Pengguna mengakses '/order/status?order_id=INVALID-ID'",
     "Sistem menampilkan halaman status dengan informasi bahwa pesanan tidak ditemukan",
     "Berhasil"),

    ("Status Pesanan",
     "Polling live status pesanan (API)",
     "Pengguna berada di halaman status, sistem melakukan polling ke '/api/orders/{id}/status'",
     "API mengembalikan JSON berisi field 'status' terkini (pending/paid/making/ready/done) dengan kode HTTP 200",
     "Berhasil"),

    ("Status Pesanan",
     "Live status pesanan tidak ditemukan",
     "Request GET ke '/api/orders/INVALID/status'",
     "API mengembalikan response JSON {'status': 'not_found'} dengan HTTP 404",
     "Berhasil"),

    ("Status Pesanan",
     "Riwayat pesanan pelanggan (API history)",
     "Pengguna mengirim GET ke '/api/orders/history?order_ids[]=SKENA-XXX'",
     "Sistem mengembalikan array JSON berisi detail pesanan (maks. 20 data) terurut dari terbaru",
     "Berhasil"),

    # ── NOTIFIKASI PEMBAYARAN MIDTRANS ────────────────────────────────────────
    ("Notifikasi Pembayaran",
     "Menerima webhook settlement dari Midtrans",
     "Midtrans mengirim POST ke '/payment/notification' dengan transaction_status = 'settlement'",
     "Sistem memperbarui status order menjadi 'paid' dan mengisi kolom paid_at dengan waktu saat ini",
     "Berhasil"),

    ("Notifikasi Pembayaran",
     "Menerima webhook capture (fraud: accept)",
     "Midtrans mengirim notifikasi dengan transaction_status = 'capture' dan fraud_status = 'accept'",
     "Sistem memperbarui status order menjadi 'paid' dan mengisi paid_at",
     "Berhasil"),

    ("Notifikasi Pembayaran",
     "Menerima webhook expire/cancel",
     "Midtrans mengirim notifikasi dengan transaction_status = 'expire' atau 'cancel'",
     "Sistem memperbarui status order menjadi 'cancelled'",
     "Berhasil"),

    ("Notifikasi Pembayaran",
     "Webhook untuk order tidak ditemukan",
     "Midtrans mengirim notifikasi dengan order_id yang tidak ada di database",
     "Sistem mengembalikan response JSON {'message': 'Order not found'} dengan HTTP 404",
     "Berhasil"),

    # ── ADMIN — LOGIN & LOGOUT ─────────────────────────────────────────────────
    ("Admin — Login",
     "Login admin dengan kredensial valid",
     "Admin mengakses '/admin/login', memasukkan email dan password yang terdaftar, lalu menekan tombol Login",
     "Sistem melakukan autentikasi, sesi dimulai, admin diarahkan ke halaman dashboard sesuai role (Admin/Kasir/Owner)",
     "Berhasil"),

    ("Admin — Login",
     "Login dengan email tidak terdaftar",
     "Admin memasukkan email yang tidak ada di database",
     "Sistem menampilkan pesan error: 'Email atau password yang Anda masukkan salah.' dan tidak membuat sesi",
     "Berhasil"),

    ("Admin — Login",
     "Login dengan password salah",
     "Admin memasukkan email yang benar tetapi password yang salah",
     "Sistem menampilkan pesan error: 'Email atau password yang Anda masukkan salah.'",
     "Berhasil"),

    ("Admin — Login",
     "Login dengan field kosong",
     "Admin menekan tombol Login tanpa mengisi email atau password",
     "Sistem menampilkan validasi: 'Email wajib diisi' dan 'Password wajib diisi'",
     "Berhasil"),

    ("Admin — Login",
     "Logout dari sistem",
     "Admin yang sudah login menekan tombol Logout",
     "Sesi admin dihapus, token diregenerasi, admin diarahkan kembali ke halaman '/admin/login'",
     "Berhasil"),

    ("Admin — Login",
     "Akses dashboard tanpa login",
     "Pengguna yang belum login mencoba mengakses '/admin/dashboard' langsung",
     "Sistem mengarahkan pengguna ke halaman '/admin/login' (middleware admin.auth aktif)",
     "Berhasil"),

    # ── ADMIN — DASHBOARD ─────────────────────────────────────────────────────
    ("Admin — Dashboard",
     "Menampilkan statistik hari ini",
     "Admin membuka halaman dashboard '/admin/'",
     "Sistem menampilkan: Total Order Hari Ini, Pendapatan Hari Ini, dan Pesanan Aktif (status: paid/making/ready)",
     "Berhasil"),

    ("Admin — Dashboard",
     "Menampilkan grafik penjualan 7 hari",
     "Admin berada di halaman dashboard",
     "Sistem menampilkan grafik bar/line penjualan 7 hari terakhir beserta label hari yang benar",
     "Berhasil"),

    ("Admin — Dashboard",
     "Menampilkan grafik penjualan 30 hari",
     "Admin memilih tab/filter '30 Hari' pada chart penjualan",
     "Sistem menampilkan grafik penjualan 30 hari terakhir dengan label tanggal format dd/mm",
     "Berhasil"),

    ("Admin — Dashboard",
     "Memperbarui status pesanan",
     "Admin menekan tombol ubah status pesanan (making → ready → done) pada panel pesanan aktif",
     "Sistem mengirim PUT ke '/admin/api/orders/{id}/status', database diperbarui, UI menampilkan status baru tanpa reload halaman",
     "Berhasil"),

    ("Admin — Dashboard",
     "Memperbarui status dengan nilai tidak valid",
     "Request PUT dikirim ke '/admin/api/orders/{id}/status' dengan status = 'invalid_status'",
     "Sistem mengembalikan response error 422 dengan pesan validasi bahwa nilai status tidak diperbolehkan",
     "Berhasil"),

    # ── ADMIN — MANAJEMEN MENU ────────────────────────────────────────────────
    ("Admin — Manajemen Menu",
     "Menampilkan daftar menu",
     "Admin membuka panel manajemen menu",
     "Sistem memuat dan menampilkan seluruh menu beserta kategori, harga, dan status ketersediaan melalui API GET '/admin/api/menus'",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Menambahkan menu baru",
     "Admin mengisi form tambah menu (nama, kategori, harga) dan menekan Simpan",
     "Sistem menyimpan data menu baru ke database, mengembalikan response JSON 200 dengan pesan 'Menu berhasil ditambahkan', menu tampil di daftar",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Menambahkan menu dengan upload gambar",
     "Admin memilih file gambar (JPEG/PNG/WEBP, maks 2MB) saat menambah menu",
     "Gambar tersimpan di storage/app/public/menus/, atribut 'image' pada record menu terisi path file",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Upload gambar dengan format tidak valid",
     "Admin mengupload file berformat .pdf atau .gif",
     "Sistem mengembalikan error validasi: 'File harus berupa gambar (jpeg, png, jpg, webp)'",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Upload gambar melebihi ukuran maksimum",
     "Admin mengupload gambar berukuran lebih dari 2MB",
     "Sistem mengembalikan error validasi: 'Ukuran gambar maksimal 2MB'",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Mengedit data menu yang sudah ada",
     "Admin memilih menu dari daftar, mengubah nama atau harga, lalu menekan Perbarui",
     "Sistem memperbarui data menu di database, gambar lama dihapus jika ada gambar baru, mengembalikan pesan 'Menu berhasil diperbarui'",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Menghapus menu",
     "Admin menekan tombol Hapus pada salah satu menu dan mengkonfirmasi",
     "Sistem menghapus record menu dari database, file gambar terkait dihapus dari storage, menu hilang dari daftar",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Menambah menu tanpa nama (validasi)",
     "Admin mengirim form tanpa mengisi field nama menu",
     "Sistem mengembalikan error validasi 422: 'Nama menu wajib diisi'",
     "Berhasil"),

    ("Admin — Manajemen Menu",
     "Menambah menu dengan category_id tidak valid",
     "Admin mengirim POST dengan category_id yang tidak ada di tabel categories",
     "Sistem mengembalikan error validasi 422: 'Kategori tidak ditemukan'",
     "Berhasil"),

    # ── ADMIN — MANAJEMEN KATEGORI ────────────────────────────────────────────
    ("Admin — Manajemen Kategori",
     "Menampilkan daftar kategori",
     "Admin membuka panel kategori",
     "Sistem memuat seluruh kategori melalui GET '/admin/api/categories' dan menampilkannya secara berurutan (sort_order)",
     "Berhasil"),

    ("Admin — Manajemen Kategori",
     "Menambahkan kategori baru",
     "Admin mengisi nama kategori dan menekan Simpan",
     "Kategori baru tersimpan di database dan muncul di daftar kategori",
     "Berhasil"),

    ("Admin — Manajemen Kategori",
     "Mengedit kategori",
     "Admin mengubah nama kategori yang sudah ada",
     "Data kategori diperbarui di database, perubahan tercermin pada daftar menu di halaman publik",
     "Berhasil"),

    ("Admin — Manajemen Kategori",
     "Menghapus kategori",
     "Admin menekan tombol Hapus pada kategori",
     "Kategori dihapus dari database. Jika kategori masih memiliki menu terkait, sistem menangani constraint dengan tepat",
     "Berhasil"),

    # ── ADMIN — MANAJEMEN MEJA ────────────────────────────────────────────────
    ("Admin — Manajemen Meja",
     "Menampilkan daftar meja",
     "Admin membuka panel manajemen meja",
     "Sistem memuat seluruh meja melalui GET '/admin/api/tables' dan menampilkannya terurut berdasarkan nomor meja",
     "Berhasil"),

    ("Admin — Manajemen Meja",
     "Menambahkan meja baru",
     "Admin mengisi nomor meja (misal: 5), nama opsional, kapasitas, lalu menekan Simpan",
     "Meja baru tersimpan di database dengan nama default 'Meja 5' jika nama dikosongkan, response HTTP 201 dikembalikan",
     "Berhasil"),

    ("Admin — Manajemen Meja",
     "Menambahkan meja dengan nomor duplikat",
     "Admin memasukkan nomor meja yang sudah ada di database",
     "Sistem mengembalikan error validasi 422: 'Nomor meja sudah digunakan'",
     "Berhasil"),

    ("Admin — Manajemen Meja",
     "Mengedit data meja",
     "Admin mengubah kapasitas atau nama meja yang sudah ada",
     "Data meja diperbarui di database, response JSON 200 dikembalikan dengan data meja terbaru",
     "Berhasil"),

    ("Admin — Manajemen Meja",
     "Menghapus meja",
     "Admin menekan tombol Hapus pada salah satu meja dan mengkonfirmasi",
     "Meja dihapus dari database, tidak muncul lagi di daftar meja",
     "Berhasil"),

    # ── ADMIN — ANALYTICS ────────────────────────────────────────────────────
    ("Admin — Analytics",
     "Menampilkan data analytics",
     "Admin mengakses GET '/admin/api/analytics/data'",
     "API mengembalikan data JSON berisi statistik penjualan, grafik, dan performa menu",
     "Berhasil"),

    ("Admin — Analytics",
     "Export laporan ke CSV",
     "Admin menekan tombol 'Export CSV'",
     "Sistem menghasilkan dan mengunduh file .csv berisi data laporan penjualan",
     "Berhasil"),

    ("Admin — Analytics",
     "Export laporan ke PDF",
     "Admin menekan tombol 'Export PDF'",
     "Sistem menghasilkan dan mengunduh file .pdf berisi laporan penjualan yang terformat",
     "Berhasil"),

    # ── ADMIN — GANTI PASSWORD ────────────────────────────────────────────────
    ("Admin — Pengaturan",
     "Ganti password berhasil",
     "Admin mengisi password lama yang benar, password baru, dan konfirmasi password baru",
     "Sistem memverifikasi password lama, menyimpan hash password baru, mengembalikan JSON {'success': true, 'message': 'Kata sandi berhasil diubah.'}",
     "Berhasil"),

    ("Admin — Pengaturan",
     "Ganti password — password lama salah",
     "Admin mengisi password lama yang tidak cocok dengan yang tersimpan di database",
     "Sistem mengembalikan response JSON {'success': false, 'message': 'Kata sandi saat ini tidak cocok.'} dengan HTTP 400",
     "Berhasil"),

    ("Admin — Pengaturan",
     "Ganti password — konfirmasi tidak cocok",
     "Admin mengisi password baru namun kolom konfirmasi diisi berbeda",
     "Sistem mengembalikan error validasi 422: 'Konfirmasi password tidak cocok'",
     "Berhasil"),

    ("Admin — Pengaturan",
     "Ganti password — password baru terlalu pendek",
     "Admin mengisi password baru dengan kurang dari 6 karakter",
     "Sistem mengembalikan error validasi 422: 'Password minimal 6 karakter'",
     "Berhasil"),

    # ── KASIR — DASHBOARD ─────────────────────────────────────────────────────
    ("Kasir — Dashboard",
     "Akses dashboard kasir dengan role yang benar",
     "Pengguna dengan role 'kasir' login dan mengakses '/kasir/'",
     "Sistem menampilkan dashboard kasir dengan daftar pesanan aktif",
     "Berhasil"),

    ("Kasir — Dashboard",
     "Akses dashboard kasir dengan role admin (bukan kasir)",
     "Pengguna dengan role 'admin' mencoba mengakses '/kasir/'",
     "Sistem menolak akses karena middleware kasir.auth aktif, pengguna diarahkan ke halaman yang sesuai atau diberi pesan error 403",
     "Berhasil"),

    ("Kasir — Dashboard",
     "Memperbarui status pesanan oleh kasir",
     "Kasir menekan tombol ubah status pesanan (making/ready/done) pada panel pesanan",
     "Sistem mengirim PUT ke '/kasir/api/orders/{id}/status', status pesanan diperbarui di database",
     "Berhasil"),

    ("Kasir — Dashboard",
     "Export laporan penjualan kasir ke CSV",
     "Kasir mengakses GET '/kasir/api/analytics/export-csv'",
     "Sistem mengunduh file CSV berisi data penjualan yang dikelola kasir",
     "Berhasil"),

    # ── OWNER — DASHBOARD ────────────────────────────────────────────────────
    ("Owner — Dashboard",
     "Akses dashboard owner",
     "Pengguna dengan role 'owner' login dan mengakses '/owner/'",
     "Sistem menampilkan dashboard owner dengan statistik bisnis dan grafik analitik",
     "Berhasil"),

    ("Owner — Dashboard",
     "Melihat data analytics owner",
     "Owner mengakses GET '/owner/api/analytics/data'",
     "API mengembalikan data JSON berisi laporan keuangan dan statistik bisnis lengkap",
     "Berhasil"),

    ("Owner — Dashboard",
     "Export laporan owner ke PDF",
     "Owner menekan tombol 'Export PDF' di halaman analytics",
     "Sistem menghasilkan dan mengunduh file PDF berisi laporan bisnis terformat",
     "Berhasil"),

    # ── KEAMANAN & OTORISASI ──────────────────────────────────────────────────
    ("Keamanan & Otorisasi",
     "Akses route admin tanpa autentikasi",
     "Pengguna yang belum login mengakses URL '/admin/api/menus'",
     "Sistem mengembalikan redirect ke halaman login atau response HTTP 401/403",
     "Berhasil"),

    ("Keamanan & Otorisasi",
     "Proteksi CSRF pada form POST",
     "Pengguna mencoba mengirim POST ke endpoint tanpa token CSRF (kecuali webhook Midtrans)",
     "Sistem menolak request dengan response HTTP 419 (CSRF Token Mismatch)",
     "Berhasil"),

    ("Keamanan & Otorisasi",
     "Webhook Midtrans tanpa CSRF",
     "Midtrans mengirim POST ke '/payment/notification' tanpa token CSRF",
     "Sistem berhasil menerima dan memproses notifikasi (route dikecualikan dari CSRF middleware)",
     "Berhasil"),
]

# ─── Write Data ──────────────────────────────────────────────────────────────
ROW_START = 3
current_row = ROW_START
no = 1

# Group rows by feature for merging
feature_groups = {}
for item in test_data:
    fitur = item[0]
    if fitur not in feature_groups:
        feature_groups[fitur] = []
    feature_groups[fitur].append(item)

for fitur_idx, (fitur, items) in enumerate(feature_groups.items()):
    start_row = current_row
    alt_fill = fitur_idx % 2 == 0

    for i, (_, desc, kondisi, hasil, status) in enumerate(items):
        row_fill_hex = LIGHT_CREAM if alt_fill else WHITE_ROW
        row_fill     = PatternFill("solid", fgColor=row_fill_hex)

        # Col A: No
        c_no = ws.cell(row=current_row, column=1, value=no)
        c_no.font      = Font(name="Calibri", size=9, bold=True, color=DARK_BROWN)
        c_no.fill      = row_fill
        c_no.alignment = Alignment(horizontal="center", vertical="center")
        c_no.border    = thin_border()

        # Col B: Fitur (will be merged later)
        c_feat = ws.cell(row=current_row, column=2, value=fitur if i == 0 else "")
        c_feat.font      = Font(name="Calibri", size=9, bold=True, color=TEXT_WHITE)
        c_feat.fill      = PatternFill("solid", fgColor=MEDIUM_BROWN)
        c_feat.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c_feat.border    = thin_border()

        # Col C: Deskripsi
        c_desc = ws.cell(row=current_row, column=3, value=desc)
        c_desc.font      = Font(name="Calibri", size=9, color=TEXT_DARK)
        c_desc.fill      = row_fill
        c_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_desc.border    = thin_border()

        # Col D: Kondisi
        c_kond = ws.cell(row=current_row, column=4, value=kondisi)
        c_kond.font      = Font(name="Calibri", size=9, color=TEXT_DARK)
        c_kond.fill      = row_fill
        c_kond.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_kond.border    = thin_border()

        # Col E: Hasil
        c_hasil = ws.cell(row=current_row, column=5, value=hasil)
        c_hasil.font      = Font(name="Calibri", size=9, color=TEXT_DARK)
        c_hasil.fill      = row_fill
        c_hasil.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_hasil.border    = thin_border()

        # Col F: Status
        status_fill = PatternFill("solid", fgColor=BERHASIL_BG if status == "Berhasil" else GAGAL_BG)
        status_color = "1A7A45" if status == "Berhasil" else "A82020"
        c_stat = ws.cell(row=current_row, column=6, value=status)
        c_stat.font      = Font(name="Calibri", size=9, bold=True, color=status_color)
        c_stat.fill      = status_fill
        c_stat.alignment = Alignment(horizontal="center", vertical="center")
        c_stat.border    = thin_border()

        ws.row_dimensions[current_row].height = 52
        current_row += 1
        no += 1

    # Merge feature column
    end_row = current_row - 1
    if end_row > start_row:
        ws.merge_cells(f"B{start_row}:B{end_row}")

# ─── Footer ──────────────────────────────────────────────────────────────────
ws.merge_cells(f"A{current_row}:F{current_row}")
footer = ws[f"A{current_row}"]
footer.value     = f"Total Kasus Uji: {no - 1}   |   Semua pengujian dilakukan pada aplikasi Skena Coffee (Laravel)   |   Tanggal: 13 Agustus 2026"
footer.font      = Font(name="Calibri", size=8, italic=True, color=ACCENT_GOLD)
footer.fill      = PatternFill("solid", fgColor=DARK_BROWN)
footer.alignment = Alignment(horizontal="center", vertical="center")
footer.border    = thin_border(DARK_BROWN)
ws.row_dimensions[current_row].height = 20

# ─── Freeze panes (header always visible) ────────────────────────────────────
ws.freeze_panes = "A3"

# ─── Print settings ──────────────────────────────────────────────────────────
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows       = "1:2"

# ─── Save ────────────────────────────────────────────────────────────────────
out_path = r"c:\laragon\www\skena-coffe\Blackbox_Testing_SkenaCoffe.xlsx"
wb.save(out_path)
print(f"[OK] File Excel berhasil dibuat: {out_path}")
print(f"     Total kasus uji: {no - 1}")
